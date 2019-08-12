from openpyxl import Workbook, load_workbook
import subs as s

def write(filename):
    lines=0
    book = Workbook()
    sheet = book.active
    title = [['Sr.No','Name',
              'Content',
              '%',
              'State',
              'Molar Wt',
              'Fr. Pt',
              'B. Pt',
              'Comments',
              'Performance']]
    for t in title:
        sheet.append(t)
        lines+=1
    for y in range(len(s.arr)): #enter n number of rows
        vals = [[y+1,
                s.arr[y].name,
                s.arr[y].cont[0][0],
                s.arr[y].cont[0][1],
                s.arr[y].form,
                s.arr[y].mWt,
                s.arr[y].fPt,
                s.arr[y].bPt,
                s.arr[y].cmnts]]
        for v in vals:
            sheet.append(v)
            lines+=1
        for q in range(len(s.arr[y].pfrm)):
            sheet.cell(row=lines,column=10+q).value=s.arr[y].pfrm[q]
        for n in range(1,len(s.arr[y].cont)):
            lines +=1
            sheet.cell(row=lines,column=3).value= s.arr[y].cont[n][0]
            sheet.cell(row=lines,column=4).value= s.arr[y].cont[n][1]    
    book.save(filename+'.xlsx')

def read(filename):
    ign = 0
    bk = load_workbook(filename+'.xlsx')
    sheet = bk.active
    r,c=2,1
    flag=1
    temp = s.substance()
    while flag==1:
        if temp.name!= 'foo':
            s.arr.append(temp)
        temp = s.substance()
        ign = sheet.cell(row=r, column=1).value
        if ign != None:
            temp.name = sheet.cell(row=r,column=2).value
            temp.cont =[]
            temp.cont.append([sheet.cell(row=r,column=3).value,sheet.cell(row=r,column=4).value])
            temp.form = sheet.cell(row=r,column=5).value
            temp.mWt = sheet.cell(row=r,column=6).value
            temp.fPt = sheet.cell(row=r,column=7).value
            temp.bPt = sheet.cell(row=r,column=8).value
            temp.cmnts = sheet.cell(row=r,column=9).value
            c=10
            temp.pfrm=[]
            while sheet.cell(row=r,column=c).value!=None:
                temp.pfrm.append(sheet.cell(row=r,column=c).value)
                c+=1
            r+=1
        else:
            if temp.name=='foo':
                flag=0
            else:
                temp.cont.append([sheet.cell(row=r,column=3).value,sheet.cell(row=r,column=4).value])
                r+=1
