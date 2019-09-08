'''
name : Atmospheric model 
date : 29/8/19
time : 06:14PM
author : tanmay
pre-req :file name with data of temp and press variaiton at diff heights.
'''
from openpyxl import load_workbook

t=[]
p=[]
h=[]

def read(filename):
    bk = load_workbook(filename+'.xlsx')
    sheet = bk.active
    r,c=3,1
    flag=1
    while True:
        h.append(int(sheet.cell(row=r, column=1).value))
        t.append(float(sheet.cell(row=r, column=2).value))
        p.append(float(sheet.cell(row=r, column=3).value)*float(sheet.cell(row=2, column=3).value))
        r+=1
        if r==19: #no of data rows
            break
        
def ret_PH(val,typ): #input value and output type expected
    for i in range(len(h)-1):
        if typ=='p': #find pressure at height 'val'
            if val>=h[i] and val<h[i+1]+1:
                tru= ((p[i+1]-p[i])/(h[i+1]-h[i])*(val-h[i]))+p[i]
                t0= ((t[i+1]-t[i])/(h[i+1]-h[i])*(val-h[i]))+t[i]
                break
        elif typ=='h':  #find height at pressure 'val'
            if val<p[i]+1 and val>=p[i+1]:
                tru= ((h[i+1]-h[i])/(p[i]-p[i+1])*(p[i]-val))+h[i]
                t0= ((t[i+1]-t[i])/(h[i+1]-h[i])*(tru-h[i]))+t[i]
                break
    return [tru,t0]


    
def find(x,t): #  input value, value expected to find
    read('Atm_model_US')
    return ret_PH(x,t)

if __name__=='__main__':
    print (find(400000,'p'))
    
    
        
            
