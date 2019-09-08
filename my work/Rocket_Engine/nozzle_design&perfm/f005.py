'''
name : optimal engine design altitude and its performance over the range of flight heigth
date : 8/9/19
time : 04:00PM
author : tanmay
'''
import rocketEngine as re
from ModAtm import find
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

p1=re.pc  #combustion chamber pressure
t1=re.tc   #combustion chamber max. temp
M=re.mol_mass    #mol mass of products grams
gam=re.gamma
F =re.thrust 
max_ht=re.alt
file_name='Eng'+str(1)+'.xlsx'
R=8.314 #kJ/kg-K

def design(height,j): # p2==p3 optimal conditions from height (in pascals)
    ig=find(height*1000,'p')
    p2=ig[0] #get pressure at height 'h'
    p_x=p2/p1
    blk1= 2* pow(gam,2) /(gam -1)
    blk2 = pow((2/(gam+1)),((gam+1)/(gam-1)))
    one_min_Prat= pow(1-pow(p_x,(gam-1)/gam),0.5)
    cf_y=pow(blk1*blk2,0.5) * one_min_Prat
    
    At=F/(p1*cf_y)
    blk3=pow(((gam+1)/(gam-1)),0.5)
    blk4=pow(((gam+1)/2),(1/(gam-1)))
    A2= At / (blk3 * blk4 * one_min_Prat * pow(p_x,(1/gam)))

    c = pow((2 * (gam/(gam-1)) * (R / M ) * t1),0.5) * one_min_Prat

    #save specs for nozzle design for specific altitude
    book = load_workbook(file_name)
    sheet = book.active
    nozz_no=2*j-1
    sheet.cell(row=1,column=nozz_no).value=p_x   #press ratio
    sheet.cell(row=2,column=nozz_no).value=At    #throat area
    sheet.cell(row=3,column=nozz_no).value=A2/At #e
    sheet.cell(row=4,column=nozz_no).value=M     #avg. mol mass
    sheet.cell(row=5,column=nozz_no).value=gam   #gamma
    sheet.cell(row=6,column=nozz_no).value=c     #exhaust velocity at optimal expansion
    sheet.cell(row=7,column=nozz_no).value=cf_y  #optimal->coeff of force
    sheet.cell(row=8,column=nozz_no).value=one_min_Prat  #for calcs
    sheet.cell(row=9,column=nozz_no).value=height
    book.save(file_name)

def noz_cf_fn(nozz_no):
    #get nozzle specs.
    col = 2 * nozz_no - 1
    book = load_workbook(file_name)
    sheet = book.active
    lcl_p_x = sheet.cell(row=1,column=col).value
    lcl_At = sheet.cell(row=2,column=col).value
    lcl_e = sheet.cell(row=3,column=col).value
    lcl_A2 = lcl_e * lcl_At
    lcl_one_min_Prat=sheet.cell(row=8,column=col).value
    lcl_hgt = (sheet.cell(row=9,column=col).value)
    print('designed for height',lcl_hgt)
    print(int(lcl_hgt*100/max_ht),'% completed')
    #calc some terms
    blk1= 2 * pow(gam,2) /(gam +1)
    blk2 = pow((2/gam+1),(gam+1/gam-1))
    cf_y=pow(blk1 * blk2,0.5) * lcl_one_min_Prat
    #calc cf as a funtion of height
    y=[]
    for i in range(lcl_hgt,max_ht):                       #height range till drag becomes negligible ie. atmostphere stops existing
        ig= find(i*1000,'p')
        y.append( cf_y + lcl_e * (lcl_p_x- ig[0]/p1))       #p3=ig[0] #ambient pressure at height 'i' 
    for i in range(len(y)):
        sheet.cell(row=i+1+lcl_hgt,column=nozz_no*2).value=y[i]
    book.save(file_name)


def plot(x,ys,e):
    for i in range(len(ys)):
        plt.plot(x,ys[i],label=e[i])
    plt.xlabel('Alt (km)')
    plt.ylabel('Cf')
    #plt.legend()
    plt.show()

    
def mk_noz(inc):
    book=Workbook(file_name)
    book.save(file_name)
    j=1
    ig=0
    while True:
        design(ig,j)
        noz_cf_fn(j)
        j+=1
        ig+=inc
        if ig>=(max_ht): #no of cell entries in excel file corresponding to each km
            break
        
    x=[]
    for i in range(1,max_ht): #length of altitude array +1
        x.append(i)
    yz=[] #list of all the y's at diff altitude
    e=[]
    book = load_workbook(file_name)
    sheet = book.active
    data= True
    col=1
    while data==True:
        y=[]
        if sheet.cell(row=1,column=2*col-1).value==None:
            data=False
        else:
            e.append(float(sheet.cell(row=2,column=2*col-1).value))
            for i in range(1,max_ht):
                if sheet.cell(row=i,column=2*col).value==None:
                    y.append(yz[0][0])
                else:
                    y.append(float(sheet.cell(row=i,column=2*col).value))
            col+=1
            yz.append(y)
    plot(x,yz,e)

    
if __name__=='__main__':
    mk_noz(5)

